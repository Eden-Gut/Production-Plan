import os
import tempfile
import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Application title and description
st.markdown(
    "<h1 style='text-align: center; color: #4B0082;'>Production Plan Data File Updater</h1>",
    unsafe_allow_html=True
)

st.markdown(
    "<p style='text-align: center; color: #4682B4;'>Upload the files, click 'Update File,' and the updated file will be available for download.</p>",
    unsafe_allow_html=True
)

# File upload section
argo_file = st.file_uploader("Upload the Argo file", type=["xlsx"])
production_plan_file = st.file_uploader("Upload the Production Plan file", type=["xlsx"])

if argo_file and production_plan_file:
    st.success("Both files have been uploaded successfully!")

    # Process files on button click
    if st.button("Update File"):
        try:
            # Read the uploaded Argo file
            argo = pd.ExcelFile(argo_file)
            raw_data = pd.read_excel(argo, sheet_name='BP 20.10.24', header=1)

            # Filter and process data from Argo
            main_df = raw_data[(raw_data["Division"] == 'PCB') &
                               (raw_data["Plan Product Type"] == 'Tool')]
            main_df['Build Product'] = main_df['Build Product'].replace({'AOI FINE HT': 'LUMINA HP', 'AOI FINE': 'LUMINA HS'})
            main_df = main_df[~main_df['Build Product'].isin(['ULTRA DIMENSION 1000','VERIWIDE-A', 'VERIFINE-A', 'DIMENSION 6',
                                                  'VERISMART-A', 'ULTRA VERIFINE-A', 'VERIWIDE', 'ULTRA DIMENSION 800 AOI', 'ULTRA DIMENSION 700 AOI', 
                                                  'APEIRON 800SBS','TORNADO', 'PRECISE HR', 'TITANIUM 900', 'CASTOR TOOL','ULTRA PERFIX 500 P', 'VERISMART' ])]
            main_df['Build Qtr - Year'] = '20' + main_df['Build Qtr'].str[2:4]
            main_df['Build Qtr - Year'] = main_df['Build Qtr - Year'].astype(int)
            main_df['Build Qtr - Quarter'] = main_df['Build Qtr'].str[5]

            current_year = datetime.now().year
            current_quarter = (datetime.now().month - 1) // 3 + 1

            # Filter current and future dates
            main_df_previous_years = main_df[main_df['Build Qtr - Year'] > current_year]
            main_df_current_year = main_df[(main_df['Build Qtr - Year'] == current_year) &
                                           (main_df['Build Qtr - Quarter'].astype(int) >= current_quarter)]
            main_df = pd.concat([main_df_previous_years, main_df_current_year])

            # Read and process Production Plan file
            production_plan = pd.ExcelFile(production_plan_file)
            product_shortcuts = pd.read_excel(production_plan, sheet_name='Product Shortcuts')
            workdays_df = pd.read_excel(production_plan, sheet_name='data for ppd')
            prev_pp = pd.read_excel(production_plan, sheet_name='PPD PCB', skiprows=16, header=0, usecols="A:T")

            main_df = pd.merge(main_df, product_shortcuts[['Build Product', 'Product']], on='Build Product', how='left')
            main_df = pd.merge(main_df, workdays_df[['Product', 'Opt', 'Ass & Mech', 'Integration', 'Pack']],
                               on='Product', how='left')
            main_df.fillna('', inplace=True)

            # Convert and clean numeric columns
            main_df['Opt WD'] = np.ceil(pd.to_numeric(main_df['Opt'], errors='coerce')).fillna(0).astype(int)
            main_df['Assy WD'] = np.ceil(pd.to_numeric(main_df['Ass & Mech'], errors='coerce')).fillna(0).astype(int)
            main_df['Int WD'] = np.ceil(pd.to_numeric(main_df['Integration'], errors='coerce')).fillna(0).astype(int)
            main_df['Pack WD'] = np.ceil(pd.to_numeric(main_df['Pack'], errors='coerce')).fillna(0).astype(int)
            main_df.drop(columns=['Opt', 'Ass & Mech', 'Integration', 'Pack', 'Build Product'], inplace=True)

            filtered_prev_pp = prev_pp[~prev_pp['Forecast ID'].isin(main_df['Forecast ID'])]
            combine_df = pd.concat([main_df, filtered_prev_pp])

            #'Product Family' is the primary sort key, 'MFG Commit Date' is the secondary sort key
            combine_df = combine_df.sort_values(by=['Product Family', 'Product','MFG Commit Date'], ascending=[True, True, True])

            combine_df = combine_df[['Forecast ID','Slot ID/UTID', 'Build Qtr', 'Forecast Product', 'Fab Name', 'MFG Commit Date', 'Product Family', 'Product',
                  'Opt Start', 'Opt WD','Opt End','Assy Start', 'Assy WD', 'Assy End', 'Int Start', 'Int WD', 'Int End',
                  'Pack Start', 'Pack WD', 'Pack End']]

            # Load the original workbook and update the data
            wb = load_workbook(production_plan_file)
            ws = wb['PPD PCB']

            # Apply styles and update the worksheet
            border_style = Border(left=Side(border_style='thin', color='000000'),
                                  right=Side(border_style='thin', color='000000'),
                                  top=Side(border_style='thin', color='000000'),
                                  bottom=Side(border_style='thin', color='000000'))
            fill_green = PatternFill(start_color='d9ecd0', end_color='d9ecd0', fill_type='solid')

            def apply_common_style(ws, df):
                start_row = 17
                start_col = 1

                # Clear existing values
                for i in range(start_row, len(df) + 18):
                    for j in range(1, 20):
                        ws.cell(row=i, column=j).value = None

                for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
                    for c_idx, v in enumerate(r, start=start_col):
                        cell = ws.cell(row=r_idx, column=c_idx)
                        cell.value = v
                        if r_idx >= start_row and c_idx <= 18:
                            cell.border = border_style
                        cell.font = Font(name='Calibri', size=10)
                        cell.alignment = Alignment(horizontal='center')
                        if r_idx == start_row:
                            cell.fill = fill_green

            apply_common_style(ws, combine_df)

            # Save the file to a temporary directory
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                temp_file_path = tmp_file.name
                wb.save(temp_file_path)

            # Indicate that the file is ready for download
            st.success("The file has been updated successfully and is ready for download!")

            # Provide a download link to the user
            with open(temp_file_path, "rb") as file:
                st.download_button(
                    label="Download Updated File",
                    data=file,
                    file_name="Updated_PPD_PCB.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        except Exception as e:
            st.error(f"An error occurred: {e}")
